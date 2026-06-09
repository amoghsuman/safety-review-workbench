"""
data_loader.py
Loads the AstroTalk NSFW case Excel file and normalises both sheets into
Python objects.

Sheet 1  — "Classification"  -> session metadata (50 rows)
Sheet 2  — "Chat Messages"   -> full message history (≈5 800 rows)

The two sheets are linked via order_id / chat_order_id.
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

# Allow running this file directly (python engine/data_loader.py)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from config import DATA_PROCESSED_DIR, EXCEL_FILE

# ---------------------------------------------------------------------------
# Column indices — zero-based, matching the actual sheet layout
# ---------------------------------------------------------------------------
# Classification: Case category | Date | order_id (formula) | chat_link |
#                 AI Reason For Flagging | Feedbacks | Action
_C_CATEGORY    = 0
_C_DATE        = 1
_C_ORDER_ID    = 2   # cell contains a REGEXEXTRACT formula; fallback is the real id
_C_CHAT_LINK   = 3
_C_AI_REASON   = 4
_C_FEEDBACK    = 5
_C_ACTION      = 6

# Chat Messages: id | timestamp_ist | chat_order_id | role | message | (extra)
_M_ID          = 0
_M_TIMESTAMP   = 1
_M_ORDER_ID    = 2
_M_ROLE        = 3
_M_MESSAGE     = 4

# Regex to pull the hardcoded fallback integer out of the IFERROR formula,
# e.g.  =IFERROR(__xludf.DUMMYFUNCTION("REGEXEXTRACT(D2,…)"),"316068745")
_ORDER_ID_RE = re.compile(r'"(\d+)"\s*\)?\s*$')

# HTML tag stripper — removes <br>, <b>, </b> etc.
_HTML_TAG_RE = re.compile(r'<[^>]+>')

# ---------------------------------------------------------------------------
# Third-party name extraction — constants
# ---------------------------------------------------------------------------

# Consultant title prefixes that appear before the name in "Hi Tarot Rithvika,"
_CONSULTANT_TITLE_PREFIXES = {
    "Tarot", "Psychic", "Astro", "Dr", "Pt", "Pandit", "Guruji",
}

# Words that look capitalised but are NOT person names
_NAME_STOPWORDS: set[str] = {
    # Template / system words
    "Hi", "Hello", "Dear", "This", "Below", "Name", "Gender", "Male", "Female",
    "DOB", "TOB", "POB", "IST", "AM", "PM", "UTC",
    "User", "USER", "Consultant", "Chat", "Astrotalk",
    # Salutations
    "Sir", "Madam", "Mam", "Bhai", "Didi", "Ji", "Bro", "Sis",
    # Greetings
    "Pranam", "Namaste", "Jai", "Radhe", "Shree", "Shri",
    # Religious / mythological (not person names in this context)
    "God", "Lord", "Bhagwan", "Allah", "Dev", "Devi", "Mata",
    "Shiva", "Krishna", "Ram", "Rama", "Hanuman", "Ganesh",
    "Durga", "Lakshmi", "Saraswati", "Vishnu", "Parshuram",
    # Months
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
    # Days
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
    # Countries / regions
    "India", "Pakistan", "Bangladesh", "Nepal", "Iran", "Singapore",
    "United", "States", "America", "England", "Canada", "Australia",
    "Europe", "Africa", "Asia", "UAE", "USA", "UK",
    # Major world cities (diaspora + dream-destination references)
    "London", "Dubai", "Paris", "Berlin", "Sydney", "Toronto",
    "Houston", "Boston", "Chicago", "Oxford", "Cambridge", "Budapest",
    "Woking", "Elmhurst", "York", "New",
    # Indian states
    "Uttar", "Pradesh", "Gujarat", "Bihar", "Rajasthan", "Punjab",
    "Madhya", "Haryana", "Maharashtra", "Karnataka", "Kerala",
    "Andhra", "Telangana", "Assam", "Odisha", "Jharkhand",
    "Uttarakhand", "Himachal", "Chhattisgarh", "Goa", "Tamil", "Nadu",
    "Bengal", "West",
    # Major Indian cities / districts
    "Delhi", "Mumbai", "Lucknow", "Saharanpur", "Sarangpur", "Ahmedabad",
    "Jaipur", "Surat", "Agra", "Noida", "Gurgaon", "Chandigarh", "Bhopal",
    "Indore", "Nagpur", "Patna", "Varanasi", "Allahabad", "Dehradun",
    "Chennai", "Kolkata", "Bangalore", "Hyderabad", "Pune",
    "Ludhiana", "Raipur", "Jalandhar", "Nashik", "Dhanbad",
    "Silchar", "Rourkela", "Kumta", "Cachar", "Kangra",
    "Mangalore", "Coimbatore",
    # Titles and honorifics
    "Mr", "Mrs", "Ms", "Miss", "Prof", "Rev",
    # Hinglish salutations / social words
    "Baba", "Guru", "Munda",
    # Astrological / consultation vocabulary
    "Kundali", "Rashi", "Nakshatra", "Lagna", "Ascendant", "Transit",
    "Horoscope", "Janam", "Patri", "Dasha", "Mahadasha",
    # Sexual / body words that sometimes appear Title-Cased in transcribed chat
    "Sexy", "Boobs", "Dick", "Cock", "Pussy", "Ass", "Nude",
    "Naked", "Wet",
    # Additional geographic noise — Middle East, South Asia, international cities
    "Shiraz", "Fars", "Tehran", "Iran", "Iraq",
    "Dubai", "Sharjah", "Riyadh",
    "Karachi", "Lahore", "Dhaka", "Colombo", "Kathmandu",
    "Chedle", "Surrey",
    # Misc capitalised words that appear in chat
    "Yes", "No", "Okay", "Ok", "Thanks", "Thank", "Please", "Sorry",
    "Kya", "Aap", "Haan", "Nahi", "Yaar", "Bete",
    # Single-letter initials that sneak through (though regex filters these)
}

# Capitalised word pattern — Title-cased words of 3+ chars
_CAPS_WORD_RE = re.compile(r'\b([A-Z][a-z]{2,})\b')

# Common English words that can appear Title-Cased mid-sentence
_COMMON_ENGLISH: set[str] = {
    "The", "And", "But", "For", "Not", "Are", "Was", "Has", "Had", "Its",
    "One", "Two", "All", "Can", "Our", "Out", "Get", "Got", "Him", "Her",
    "His", "She", "They", "You", "Your", "Their", "That", "With", "From",
    "Just", "Like", "Also", "More", "Some", "Have", "Been", "Will", "When",
    "What", "Where", "How", "Who", "Why", "Which", "Then", "Than", "Into",
    "Over", "After", "About", "Again", "Always", "Already", "Around",
    "Because", "Between", "During", "Before", "Without",
    # Common verbs Title-Cased in Hinglish sentences
    "Stop", "Come", "Take", "Call", "Tell", "Know", "Want", "Need", "Give",
    "See", "Look", "Talk", "Work", "Make", "Going", "Come", "Back", "Keep",
    # Common adjectives / adverbs
    "True", "Good", "Nice", "Same", "Real", "Very", "Much", "Only", "Even",
    "Last", "Next", "Both", "Most", "Many", "Long", "Sure", "Late", "Early",
    "Happy", "Best", "High", "Full", "Free", "Hot", "Big", "Small", "Little",
    # Common Hinglish words that capitalise in transcribed chat
    "Aaj", "Aap", "Aur", "Bhi", "Dnt", "Han", "Hain", "Hein", "Hun",
    "Kal", "Kiya", "Koi", "Main", "Mera", "Meri", "Nahi", "Nah",
    "Raha", "Rahi", "Sahi", "Toh", "Woh", "Yaar",
    # Hinglish words that look like names but are not
    "Dan",    # Hinglish particle ("Dan se" = "then/from that point")
    "Janna",  # Hinglish verb ("Janna hai" = "need to know")
    "Batao", "Payegi", "Abhi", "Hota", "Karne", "Aisa",
    "Theek", "Zindagi", "Zyada", "Evening",
    # Typos / garbled words that slip through
    "Incan",    # garbled word, "as max Incan"
    "Harding",  # typo of "harming" ("I'm Harding myself")
    "Dard", "Laal", "Dena", "Daga", "Kant", "Kohi", "Called", "Tommorow",
}


def _is_common_english(word: str) -> bool:
    """Return True if the word is a known common English / Hinglish non-name."""
    return word in _COMMON_ENGLISH


def _strip_html(text: str) -> str:
    """Remove HTML tags and normalise whitespace."""
    if not text:
        return ""
    return _HTML_TAG_RE.sub(" ", text).strip()


def _extract_order_id(cell_value: Any) -> int | None:
    """
    The order_id column holds either:
      • A cached evaluated value — plain string/int like '316068745' or 316068745.0
        (when openpyxl data_only=True and the workbook has cached calc results)
      • A raw IFERROR/REGEXEXTRACT formula string — the fallback integer is
        embedded in double-quotes at the end of the formula.

    We handle both forms.
    """
    if cell_value is None:
        return None
    # Already a number (int or float)
    if isinstance(cell_value, (int, float)):
        return int(cell_value)
    text = str(cell_value).strip()
    # Plain numeric string — the cached evaluated result
    if text.isdigit():
        return int(text)
    # Formula fallback: ="…","316068745")
    m = _ORDER_ID_RE.search(text)
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# Public data classes (plain dicts — lightweight, JSON-serialisable)
# ---------------------------------------------------------------------------
# Session schema
# {
#   "order_id":       int,
#   "date":           str,
#   "category":       str,
#   "ai_reason":      str,
#   "human_feedback": str,
#   "action":         str,
#   "chat_link":      str,
#   "messages":       [ MessageDict, … ]
# }

# Message schema
# {
#   "message_id": int,
#   "timestamp":  str,
#   "role":       str,   # "USER" | "CONSULTANT"
#   "message":    str
# }


class DataLoader:
    """
    Reads NSFW_Cases_Categorization.xlsx and exposes clean session objects.

    Usage
    -----
    loader = DataLoader()           # uses EXCEL_FILE from config.py
    sessions = loader.load()        # triggers parsing + prints summary
    s = loader.get_session(316068745)
    loader.save_processed()
    """

    VALID_CATEGORIES = {"Explicit", "Borderline", "Moderate", "False Positives"}

    def __init__(self, file_path: Path | str | None = None) -> None:
        self.file_path: Path = Path(file_path) if file_path else EXCEL_FILE
        self._sessions: dict[int, dict[str, Any]] = {}   # order_id -> session
        self._loaded = False
        self.dedup_stats: dict[int, int] = {}             # order_id -> removed count

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def load(self) -> list[dict[str, Any]]:
        """
        Parse the Excel file, link sheets, print a summary, and return
        all session objects sorted by date.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(
                f"Excel file not found: {self.file_path}\n"
                "Place NSFW_Cases_Categorization.xlsx in data/raw/ and retry."
            )

        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)

        metadata   = self._parse_classification(wb["Classification"])
        messages   = self._parse_chat_messages(wb["Chat Messages"])

        # Link messages into metadata and extract name fields
        for order_id, meta in metadata.items():
            meta["messages"] = messages.get(order_id, [])
            user_name, consultant_name, third_party_names = self._extract_names(
                meta["messages"]
            )
            meta["user_name"]          = user_name
            meta["consultant_name"]    = consultant_name
            meta["third_party_names"]  = third_party_names

        self._sessions = metadata
        self._loaded   = True
        self._print_summary()
        return self.get_all_sessions()

    def get_session(self, order_id: int) -> dict[str, Any]:
        """Return a single session dict; raises KeyError if not found."""
        self._require_loaded()
        return self._sessions[order_id]

    def get_all_sessions(self) -> list[dict[str, Any]]:
        """Return all 50 session dicts, sorted by date ascending."""
        self._require_loaded()
        return sorted(self._sessions.values(), key=lambda s: s["date"])

    def get_sessions_by_category(self, category: str) -> list[dict[str, Any]]:
        """
        Return sessions matching the given category label.
        Valid values: 'Explicit', 'Borderline', 'Moderate', 'False Positives'.
        """
        self._require_loaded()
        if category not in self.VALID_CATEGORIES:
            raise ValueError(
                f"Unknown category {category!r}. "
                f"Valid options: {sorted(self.VALID_CATEGORIES)}"
            )
        return [s for s in self._sessions.values() if s["category"] == category]

    def save_processed(self, output_name: str = "sessions.json") -> Path:
        """Write parsed sessions to data/processed/<output_name>."""
        self._require_loaded()
        DATA_PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        out = DATA_PROCESSED_DIR / output_name
        with open(out, "w", encoding="utf-8") as fh:
            json.dump(self.get_all_sessions(), fh, ensure_ascii=False, indent=2)
        print(f"[DataLoader] Saved {len(self._sessions)} sessions -> {out}")
        return out

    # ------------------------------------------------------------------
    # Private — sheet parsers
    # ------------------------------------------------------------------

    def _parse_classification(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> dict[int, dict[str, Any]]:
        """
        Parse the "Classification" sheet.
        Skips the header row and any fully-empty trailing rows.
        Returns a dict keyed by order_id.
        """
        metadata: dict[int, dict[str, Any]] = {}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue  # skip header

            # Skip entirely empty rows (trailing blank rows in the sheet)
            if all(v is None for v in row):
                continue

            order_id = _extract_order_id(row[_C_ORDER_ID])
            if order_id is None:
                continue  # can't link without an id

            metadata[order_id] = {
                "order_id":       order_id,
                "date":           self._clean_str(row[_C_DATE]),
                "category":       self._clean_str(row[_C_CATEGORY]),
                "ai_reason":      self._clean_str(row[_C_AI_REASON]),
                "human_feedback": self._clean_str(row[_C_FEEDBACK]),
                "action":         self._clean_str(row[_C_ACTION]),
                "chat_link":      self._clean_str(row[_C_CHAT_LINK]),
                "messages":       [],   # filled in after _parse_chat_messages
            }

        return metadata

    def _parse_chat_messages(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> dict[int, list[dict[str, Any]]]:
        """
        Parse the "Chat Messages" sheet, then deduplicate consecutive messages
        within each session.

        Returns a dict: chat_order_id -> [message_dict, …] (time-ordered,
        duplicates removed).
        """
        messages: dict[int, list[dict[str, Any]]] = defaultdict(list)

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue  # skip header
            if all(v is None for v in row):
                continue

            order_id = int(row[_M_ORDER_ID]) if row[_M_ORDER_ID] else None
            if order_id is None:
                continue

            msg: dict[str, Any] = {
                "message_id": int(row[_M_ID]) if row[_M_ID] else None,
                "timestamp":  self._clean_str(row[_M_TIMESTAMP]),
                "role":       self._clean_str(row[_M_ROLE]).upper(),
                "message":    _strip_html(self._clean_str(row[_M_MESSAGE])),
            }
            messages[order_id].append(msg)

        # Deduplicate within each session and record stats
        result: dict[int, list[dict[str, Any]]] = {}
        for order_id, msgs in messages.items():
            deduped, removed = self._dedup_messages(msgs)
            result[order_id]            = deduped
            self.dedup_stats[order_id]  = removed

        return result

    def _dedup_messages(
        self, messages: list[dict[str, Any]]
    ) -> tuple[list[dict[str, Any]], int]:
        """
        Remove consecutive duplicate messages from a session's message list.

        Two messages are considered duplicates when ALL of:
          - same role
          - same message text (whitespace-normalised)
          - timestamps within 5 seconds of each other

        The first occurrence is kept; subsequent duplicates are dropped.
        Returns (deduplicated_list, number_removed).
        """
        if not messages:
            return messages, 0

        kept: list[dict[str, Any]] = [messages[0]]
        removed = 0

        for current in messages[1:]:
            prev = kept[-1]

            if (
                current["role"]    == prev["role"]
                and current["message"].strip() == prev["message"].strip()
                and self._within_seconds(current["timestamp"], prev["timestamp"], 5)
            ):
                removed += 1
            else:
                kept.append(current)

        return kept, removed

    @staticmethod
    def _within_seconds(ts1: str, ts2: str, threshold: int) -> bool:
        """
        Return True if two ISO-8601 timestamp strings are within
        *threshold* seconds of each other. Returns True on any parse
        failure so that malformed timestamps never cause data loss.
        """
        try:
            t1 = datetime.fromisoformat(ts1.replace("Z", "+00:00"))
            t2 = datetime.fromisoformat(ts2.replace("Z", "+00:00"))
            return abs((t1 - t2).total_seconds()) <= threshold
        except (ValueError, AttributeError):
            return True

    # ------------------------------------------------------------------
    # Private — name extraction
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_names(
        messages: list[dict[str, Any]],
    ) -> tuple[str, str, list[str]]:
        """
        Derive three name fields from the session's message list.

        Returns
        -------
        user_name        : str   — from "Name: X" in the first USER message,
                                   or "" if redacted / absent
        consultant_name  : str   — from "Hi [Name]," in the first USER message,
                                   with title prefixes stripped
        third_party_names: list  — capitalised proper nouns found in USER
                                   messages that are neither the user's name,
                                   the consultant's name, nor a known stopword /
                                   place name.  Scans ALL USER messages so that
                                   names introduced later (e.g. "what about
                                   Priyanka?") are captured.
        """
        user_msgs = [m for m in messages if m["role"] == "USER"]
        if not user_msgs:
            return "", "", []

        first_msg = user_msgs[0]["message"]

        # ---- user_name -----------------------------------------------
        user_name = ""
        name_match = re.search(r'\bName\s*:\s*([^\s,\n]+)', first_msg)
        if name_match:
            candidate = name_match.group(1).strip()
            # Skip the redaction placeholder "USER"
            if candidate.upper() != "USER":
                user_name = candidate

        # ---- consultant_name -----------------------------------------
        # Pattern: "Hi [Title?] FirstName[...],"
        # e.g.  "Hi Baani,"  or  "Hi Tarot Rithvika,"  or  "Hi Psychic Zahira,"
        consultant_name = ""
        hi_match = re.match(r'Hi\s+([\w\s]+?)(?:\s*,|\.|\s+Below)', first_msg)
        if hi_match:
            parts = hi_match.group(1).strip().split()
            # Drop known title prefixes; take the last remaining word as the name
            name_parts = [p for p in parts if p not in _CONSULTANT_TITLE_PREFIXES]
            consultant_name = name_parts[-1] if name_parts else (parts[-1] if parts else "")

        # ---- third_party_names ---------------------------------------
        # Build the per-session exclusion set
        excluded = _NAME_STOPWORDS | {
            user_name,
            consultant_name,
            # Also exclude all words from the title portion of the greeting
            *(hi_match.group(1).split() if hi_match else []),
        }
        excluded.discard("")   # don't exclude the empty string

        seen: dict[str, int] = {}   # name -> occurrence count

        for m in user_msgs:
            for word in _CAPS_WORD_RE.findall(m["message"]):
                if word not in excluded and not _is_common_english(word):
                    seen[word] = seen.get(word, 0) + 1

        # Only keep names that appear at least twice — single occurrences are
        # usually capitalised common words, not person names.
        # Exception: if fewer than 10 unique candidates, keep singletons too
        # so short sessions still yield results.
        min_freq = 2 if len(seen) >= 10 else 1
        third_party_names = [
            name for name, cnt in sorted(seen.items(), key=lambda x: -x[1])
            if cnt >= min_freq
        ]

        return user_name, consultant_name, third_party_names

    # ------------------------------------------------------------------
    # Private — utilities
    # ------------------------------------------------------------------

    @staticmethod
    def _clean_str(value: Any) -> str:
        """Coerce a cell value to a stripped string; return '' for None."""
        if value is None:
            return ""
        return str(value).strip()

    def _require_loaded(self) -> None:
        if not self._loaded:
            raise RuntimeError("Call loader.load() before accessing sessions.")

    # ------------------------------------------------------------------
    # Private — summary printer
    # ------------------------------------------------------------------

    def _print_summary(self) -> None:
        sessions  = list(self._sessions.values())

        total     = len(sessions)
        cat_count = Counter(s["category"] for s in sessions)
        msg_total = sum(len(s["messages"]) for s in sessions)

        dates    = [s["date"] for s in sessions if s["date"]]
        date_min = min(dates) if dates else "N/A"
        date_max = max(dates) if dates else "N/A"

        action_count  = Counter(s["action"] for s in sessions)
        total_removed = sum(self.dedup_stats.values())

        print()
        print("=" * 56)
        print("  AstroTalk NSFW DataLoader - Session Summary")
        print("=" * 56)
        print(f"  Total sessions  : {total}")
        print(f"  Total messages  : {msg_total}  (after dedup)")
        print(f"  Duplicates rmvd : {total_removed}")
        print(f"  Date range      : {date_min[:10]}  to  {date_max[:10]}")
        print()
        print("  Sessions by category:")
        for cat in sorted(self.VALID_CATEGORIES):
            n = cat_count.get(cat, 0)
            bar = "#" * n
            print(f"    {cat:<20} {n:>3}  {bar}")
        print()
        print("  Action breakdown:")
        for action, n in sorted(action_count.items()):
            print(f"    {action:<20} {n:>3}")
        print()
        print("  Avg messages / session :", round(msg_total / total, 1) if total else 0)
        print("=" * 56)
        print()


# ---------------------------------------------------------------------------
# Quick self-test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    loader   = DataLoader()
    sessions = loader.load()

    # ---- third-party name extraction report ----
    print("Third-party name extraction — three target sessions")
    print("=" * 56)
    for oid in [296029912, 294055364, 296503802]:
        s = loader.get_session(oid)
        print(f"\n  order_id         : {oid}")
        print(f"  category         : {s['category']}")
        print(f"  human_feedback   : {s['human_feedback']}")
        print(f"  user_name        : {s['user_name']!r}")
        print(f"  consultant_name  : {s['consultant_name']!r}")
        print(f"  third_party_names: {s['third_party_names']}")

    # ---- full dataset overview ----
    print(f"\n{'='*56}")
    print("Sessions with non-empty third_party_names:")
    for s in sorted(sessions, key=lambda x: x["order_id"]):
        if s["third_party_names"]:
            print(f"  {s['order_id']}  ({s['category']:<20})  {s['third_party_names'][:5]}")

    # ---- save updated JSON ----
    out = loader.save_processed()
    print(f"\nsave_processed() -> {out}")
